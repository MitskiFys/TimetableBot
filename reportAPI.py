from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
import datetime
from mysqlAPI import DBConnect
from string import Template

class DeltaTemplate(Template):
    delimiter = "%"

def strfdelta(tdelta, fmt):
    d = {"D": tdelta.days}
    hours, rem = divmod(tdelta.seconds, 3600)
    minutes, seconds = divmod(rem, 60)
    d["H"] = '{:02d}'.format(hours)
    d["M"] = '{:02d}'.format(minutes)
    d["S"] = '{:02d}'.format(seconds)
    t = DeltaTemplate(fmt)
    return t.substitute(**d)

wb = Workbook()

ws = wb.active

ws["A1"] = "Дата"
ws["B1"] = "Время начала"
ws["C1"] = "Время ухода"
ws["D1"] = "Обед"
ws["E1"] = "Время работы"
ws["G1"] = "Больничный (в днях)"
ws["G4"] = "Отпуск (в днях)"
ws["I1"] = "Всего отработано (часов)"
ws["I4"] = "Всего соц часов"
ws["I7"] = "Всего часов"

connect = DBConnect()
result = connect.getHoursAndTypeInSpecifiedMonth(338986739, 9)

rowIter = 2


for id, date, type in result:
    if type == 0:
        ws.cell(row=rowIter, column=1, value=date.strftime('%d.%m.%Y'))
        ws.cell(row=rowIter, column=2, value=date.strftime('%H:%M')).number_format = 'HH:MM'
    elif type == 1:
        ws.cell(row=rowIter, column=3, value=date.strftime('%H:%M')).number_format = 'HH:MM'
        dinnerTime = connect.getDinnerTime(id)
        ws.cell(row=rowIter, column=4, value=strfdelta(datetime.timedelta(minutes=dinnerTime), '%H:%M')).number_format = 'HH:MM'
        startTimeCell = ws.cell(row=rowIter, column=2).coordinate
        endTimeCell = ws.cell(row=rowIter, column=3).coordinate
        dinnerTime = ws.cell(row=rowIter, column=4).coordinate
        ws.cell(row=rowIter, column=5, value=f"={endTimeCell}-{startTimeCell}-{dinnerTime}").number_format = 'HH:MM'
        rowIter += 1

lastCell = ws.cell(row=rowIter - 1, column=5).coordinate
ws["I2"] = f"=СУММ(E2:{lastCell})"
ws["I2"].number_format = 'HH:MM'
    

cell_range = ws['A1':'E1']
#cell_range.alignment = Alignment(horizontal="center", vertical="center")

for row in cell_range:
    for cell in row:
        cell.alignment = Alignment(horizontal="center", vertical="center")

dim_holder = DimensionHolder(worksheet=ws)
for col in range(ws.min_column, ws.max_column + 1):
    dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=15)
ws.column_dimensions = dim_holder


wb.save("test.xlsx")